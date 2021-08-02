from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from objects import Device
from objects import Ticket

# Using an excel sheet with special exporting - see Ryder.
# Not all devices of the given model may be found in Asset Tiger

dell_tickets_file = "C:\\Users\\hagand\\Downloads\\Dell_TechDirect_Tickets.xlsx"
asset_tiger_file = "C:\\Users\\hagand\\Downloads\\AssetTiger_AllDevices_Warranty.xlsx"

dell_tickets_workbook = load_workbook(dell_tickets_file)
dell_tickets_sheet = dell_tickets_workbook["Sheet1"]

asset_tiger_workbook = load_workbook(asset_tiger_file)
asset_tiger_sheet = asset_tiger_workbook["Export"]

search_tag = ""
search_model = "OPTIPLEX 7460 AIO"
search_purchase_date = ""
search_warranty = ""
bold_font = Font(bold=True)

def read_data():
  # Read in Asset Tiger
  devices_dict = dict()
  for idx, cell in enumerate(asset_tiger_sheet['A']):
    if idx != 0:
      st = cell.value
      model = asset_tiger_sheet['D' + str(idx + 1)].value
      warranty = asset_tiger_sheet['C' + str(idx + 1)].value
      devices_dict[st] = Device(st, model, None, warranty)


  # Read in Dell Tickets
  for idx, cell in enumerate(dell_tickets_sheet['A']):
    if idx != 0:
      work_order = cell.value
      device_tag = dell_tickets_sheet['C' + str(idx)].value
      status = dell_tickets_sheet['B' + str(idx)].value
      problem = dell_tickets_sheet['E' + str(idx)].value
      date_created = dell_tickets_sheet['F' + str(idx)].value
      ticket = Ticket(work_order, device_tag, status, problem, date_created)

      device = devices_dict.get(device_tag)
      if device != None: 
        device.tickets.append(ticket)
  return devices_dict

def filter_data(devices_dict):
  filtered_devices = devices_dict
  filtered_devices = {k:v for (k, v) in filtered_devices.items() if search_tag in v.model}
  filtered_devices = {k:v for (k, v) in filtered_devices.items() if search_model in v.model}
  filtered_devices = {k:v for (k, v) in filtered_devices.items() if search_purchase_date in v.model}
  filtered_devices = {k:v for (k, v) in filtered_devices.items() if search_warranty in v.model}
  return filtered_devices

def format_date(raw_date):
  if raw_date is None:
    return raw_date
  proper_date = str(raw_date).split()[0].split("-")
  proper_date = proper_date[1] + "-" + proper_date[2] + "-" + proper_date[0]
  return proper_date

def format_date_yrmo(raw_date):
  if raw_date is None:
    return raw_date
  proper_date = str(raw_date).split()[0].split("-")
  proper_date = proper_date[0] + "-" + proper_date[1]
  return proper_date

devices_dict = read_data()
filtered_devices = filter_data(devices_dict) 

wb = Workbook()
dest_filename = 'asset_list.xlsx'
ws0 = wb.active
ws0.title = "Overview"
ws1 = wb.create_sheet()
ws1.title = "General Asset Info"
ws2 = wb.create_sheet()
ws2.title = "Ticket Statistics"
ws3 = wb.create_sheet()
ws3.title = "Warranty to Ticket Ratio"

def create_overview_assessment(ws):
  # ======= Search Criteria
  ws['A1'] = "Search Criteria"
  ws['A2'] = "Service Tag"
  if search_tag == "":
    ws['B2'] = "N/A"
  else:
    ws['B2'] = search_tag

  ws['A3'] = "Model"
  if search_model == "":
    ws['B3'] = "N/A"
  else:
    ws['B3'] = search_model

  ws['A4'] = "Purchase Date"
  if search_purchase_date == "":
    ws['B4'] = "N/A"
  else:
    ws['B4'] = search_purchase_date

  ws['A5'] = "Warranty Date"
  if search_warranty == "":
    ws['B5'] = "N/A"
  else:
    ws['B5'] = search_warranty

  # ======= Result Info
  ws['A7'] = "Device Count"
  ws['B7'] = len(filtered_devices)

  expired = {k:v for (k, v) in filtered_devices.items() if v.hasExpired()}
  ws['A8'] = "Out of Warranty"
  ws['B8'] = len(expired)

  has_tickets = {k:v for (k, v) in filtered_devices.items() if len(v.tickets) > 0}
  ws['A9'] = "Have Had Tickets"
  ws['B9'] = len(has_tickets)

  have_multiple_tickets = {k:v for (k, v) in has_tickets.items() if len(v.tickets) > 1}
  ws['A10'] = "Have Had Multiple Tickets"
  ws['B10'] = len(have_multiple_tickets)

  ws['A12'] = "Ticket Categories"

  mobo_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "MOBO"}
  ws['A13'] = "Motherboard"
  ws['B13'] = len(mobo_issues)

  hdd_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "HDD"}
  ws['A14'] = "Hard Drive"
  ws['B14'] = len(hdd_issues)
  
  lcd_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "LCD"}
  ws['A15'] = "LCD/Screen"
  ws['B15'] = len(lcd_issues)

  psu_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "PSU"}
  ws['A16'] = "Power Supply"
  ws['B16'] = len(psu_issues)

  ram_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "RAM"}
  ws['A17'] = "RAM/Memory"
  ws['B17'] = len(ram_issues)

  other_issues = {k:v for (k, v) in has_tickets.items() if v.tickets[0].category == "OTHER"}
  ws['A18'] = "Other"
  ws['B18'] = len(other_issues)
  ws['A19'] = "Other Issues Include:"
  idx = 19
  for device in other_issues:
    tickets = other_issues[device].tickets
    unique_tickets = []
    for ticket in tickets:
      if ticket.problem not in unique_tickets:
        unique_tickets.append(ticket.problem)
        ws['B' + str(idx)] = ticket.problem
        idx = idx + 1

def create_gen_asset_info_ws(ws):
  # Set Headers
  ws['A1'] = "Purchased"
  ws['B1'] = "Model"
  ws['C1'] = "Service Tag"
  ws['D1'] = "Warranty Expiration"
  ws['E1'] = "Tickets"
  ws['F1'] = "Ticket Categories"

  idx = 2
  for asset in filtered_devices:
    device = filtered_devices[asset]
    p_date = ""
    we_date_config = str(device.warranty_expiration).split()[0].split("-")
    we_date = we_date_config[1] + "-" + we_date_config[2] + "-" + we_date_config[0]
    ws['A' + str(idx)] = format_date(device.purchase_date)
    ws['B' + str(idx)] = device.model
    ws['C' + str(idx)] = device.service_tag
    ws['D' + str(idx)] = format_date(device.warranty_expiration)
    ws['E' + str(idx)] = len(device.tickets)
    problem_types = ""
    for ticket in device.tickets:
      cat = ticket.category
      if cat not in problem_types:
        problem_types = problem_types + " " + cat
    ws['F' + str(idx)] = problem_types
    idx = idx + 1

  set_col_width(ws)

def create_ticket_view_ws(ws):
  has_tickets = {k:v for (k, v) in filtered_devices.items() if len(v.tickets) > 0}
  # Set Headers
  ws['A1'] = "Service Tag"
  ws['B1'] = "Warranty Expiration Date"
  ws['C1'] = "Ticket Created"
  ws['D1'] = "Category"
  ws['E1'] = "Problem"
  ws['F1'] = "Status"

  idx = 2
  for asset in has_tickets:
    device = has_tickets[asset]
    tickets = device.tickets
    for ticket in tickets:
      ws['A' + str(idx)] = device.service_tag
      ws['B' + str(idx)] = format_date(device.warranty_expiration)
      ws['C' + str(idx)] = format_date(ticket.date_created)
      ws['D' + str(idx)] = ticket.category
      ws['E' + str(idx)] = ticket.problem
      ws['F' + str(idx)] = ticket.status
      idx = idx + 1

def create_warranty_ticket_ratios(ws):
  # Key = Month-Year, Value = [Expiring, Tickets]
  ws['B1'] = "Devices Expiring"
  ws['C1'] = "Tickets Created"

  ratio = dict()
  for asset in filtered_devices:
    device = filtered_devices[asset]
    # Device Updates
    warranty_date = format_date_yrmo(device.warranty_expiration)
    if warranty_date in ratio:
      ratio_date = ratio[warranty_date]
      ratio_date = [ratio_date[0] + 1, ratio_date[1]]
      ratio[warranty_date] = ratio_date
    else:
      ratio[warranty_date] = [1, 0]

    # Ticket Updates
    tickets = device.tickets
    for ticket in tickets:
      ticket_date = format_date_yrmo(ticket.date_created)
      if ticket_date in ratio:
        ratio_date = ratio[ticket_date]
        ratio_date = [ratio_date[0], ratio_date[1] + 1]
        ratio[ticket_date] = ratio_date
      else:
        ratio[ticket_date] = [0, 1]

  idx = 2
  for key in sorted (ratio.keys()):
    date = ratio[key]
    display_date = key.split("-")
    ws['A' + str(idx)] = display_date[1] + "-" + display_date[0]
    ws['B' + str(idx)] = date[0]
    ws['C' + str(idx)] = date[1]
    idx = idx + 1

def set_col_width(ws):
  col_widths = []
  for row in ws:
    for i, cell in enumerate(row):
      cell_val = str(cell.value)
      if cell_val != None:
        if len(col_widths) > i:
          if len(cell_val) > col_widths[i]:
            col_widths[i] = len(cell_val)
        else:
          col_widths += [len(cell_val)]

  for i, col_width in enumerate(col_widths):
    # +1 for wiggle room
    ws.column_dimensions[get_column_letter(i + 1)].width = col_width + 1

def style_workbook(ws_arr):
  for row in ws_arr[0]:
    row[0].font = bold_font

  for cell in ws_arr[1][1]:
    cell.font = bold_font
  for cell in ws_arr[2][1]:
    cell.font = bold_font
  for cell in ws_arr[3][1]:
    cell.font = bold_font

  set_col_width(ws_arr[0])
  set_col_width(ws_arr[1])
  set_col_width(ws_arr[2])
  set_col_width(ws_arr[3])

create_overview_assessment(ws0)
create_gen_asset_info_ws(ws1)
create_ticket_view_ws(ws2)
create_warranty_ticket_ratios(ws3)
style_workbook([ws0, ws1, ws2, ws3])
wb.save(filename = dest_filename)